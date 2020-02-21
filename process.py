import openpyxl as opxl

def data(file_path):
    #Creates the working book and working sheet objects as attributes of Xls().
    working_book = opxl.load_workbook(file_path)

    return working_book

def column_index(file_path, field_name):

    sheet = data(file_path).active

    #Iteractively search for the column name provided using <str.casefold()> to avoid case sensitive errors.
    for i in range(1, ((sheet.max_column) + 1)):
        if str(sheet.cell(1, i).value).casefold() == str(field_name).casefold():

            return sheet.cell(1, i).col_idx

def add_field(file_path, field_name):

    wb = data(file_path)
    sheet = wb.active
    sheet.cell(1, (sheet.max_column + 1)).value = field_name
    wb.save(file_path)

def write_value(file_path, field_name, row_number, value):

    sheet = data(file_path)
    sheet.active.cell(row_number, column_index(file_path, field_name)).value = value
    sheet.save(file_path)

